#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Adaptador de libros de Excel
============================
Entorno: venv del proyecto.

Por qué existe y por qué aquí. El consultor pide los resultados espaciales en
Excel y no en ráster: un libro se abre, se revisa y se anexa, y un ráster exige
un SIG para mirarlo. Todo lo que toca openpyxl vive en este archivo, aislado de
los módulos que calculan (CLAUDE.md, sección 2).

Sigue la misma separación que 'graficos.py', 'dss.py' y 'docx_plantilla.py':

    src/comun/     solo librería estándar    ambos entornos
    src/excel.py   depende de openpyxl        módulos de análisis

EL CSV SIGUE SIENDO EL PRODUCTO DE INTERCAMBIO. Cada hoja del libro procede de
una tabla que también se escribe en CSV, porque es lo que el siguiente módulo
lee y lo que sobrevive a un cambio de versión de Excel. El libro es para la
persona que revisa, no para la cadena.
"""

from __future__ import annotations

from pathlib import Path
from typing import Any, Sequence

__all__ = ["ErrorExcel", "escribir_libro"]

# Excel corta los nombres de hoja en 31 caracteres y prohíbe unos cuantos
# signos. Recortar en silencio produciría dos hojas con el mismo nombre y la
# segunda sobrescribiría a la primera.
LIMITE_NOMBRE = 31
PROHIBIDOS = set(r"[]:*?/\\")


class ErrorExcel(RuntimeError):
    """Falla al escribir un libro, que el módulo debe reportar."""


def nombre_de_hoja(propuesto: str, usados: Sequence[str] = ()) -> str:
    """
    Ajusta un nombre a lo que Excel admite, sin colisionar con los ya usados.

    Si el recorte produce un nombre repetido se le añade un sufijo numérico: dos
    hojas homónimas no dan error, la segunda sustituye a la primera y la tabla
    desaparece sin que nada lo señale.
    """
    limpio = "".join("_" if c in PROHIBIDOS else c for c in propuesto).strip()
    limpio = (limpio or "Hoja")[:LIMITE_NOMBRE]
    if limpio not in usados:
        return limpio
    for sufijo in range(2, 100):
        marca = f"_{sufijo}"
        candidato = limpio[:LIMITE_NOMBRE - len(marca)] + marca
        if candidato not in usados:
            return candidato
    raise ErrorExcel(f"no se pudo dar nombre unico a la hoja {propuesto!r}.")


def escribir_libro(
    destino: Path, hojas: Sequence[tuple[str, Sequence[dict[str, Any]]]],
    congelar_encabezado: bool = True,
) -> dict[str, Any]:
    """
    Escribe un libro con una hoja por tabla, en el orden recibido.

    Las columnas de cada hoja son la unión de las claves de sus filas, en el
    orden en que aparecen: así una fila con un campo de más no descuadra la
    tabla ni se pierde.

    Excepciones
    -----------
    ErrorExcel
        Si falta la librería o si la escritura falla.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font
        from openpyxl.utils import get_column_letter
    except ImportError as error:  # pragma: no cover - depende del entorno
        raise ErrorExcel(
            "no esta instalado 'openpyxl', que es lo que escribe el libro. "
            "Instalarlo en el venv del proyecto.") from error

    libro = Workbook()
    libro.remove(libro.active)
    escritas: list[dict[str, Any]] = []
    usados: list[str] = []

    for propuesto, filas in hojas:
        filas = list(filas)
        nombre = nombre_de_hoja(propuesto, usados)
        usados.append(nombre)
        hoja = libro.create_sheet(nombre)

        columnas: list[str] = []
        for fila in filas:
            for clave in fila:
                if clave not in columnas:
                    columnas.append(clave)
        if columnas:
            hoja.append(columnas)
            for celda in hoja[1]:
                celda.font = Font(bold=True)
            for fila in filas:
                hoja.append([fila.get(c) for c in columnas])
            if congelar_encabezado:
                hoja.freeze_panes = "A2"
            for indice, clave in enumerate(columnas, start=1):
                ancho = max([len(str(clave))]
                            + [len(str(f.get(clave, ""))) for f in filas[:200]])
                hoja.column_dimensions[get_column_letter(indice)].width = min(
                    max(ancho + 2, 10), 40)
        escritas.append({"hoja": nombre, "filas": len(filas),
                         "columnas": len(columnas)})

    if not escritas:
        raise ErrorExcel("no se recibio ninguna hoja que escribir.")

    destino = Path(destino)
    destino.parent.mkdir(parents=True, exist_ok=True)
    try:
        libro.save(str(destino))
    except OSError as error:
        raise ErrorExcel(
            f"no se pudo escribir {destino.name}: {error}. Si el libro esta "
            "abierto en Excel, cerrarlo y repetir.") from error

    return {"archivo": str(destino), "hojas": escritas,
            "kb": round(destino.stat().st_size / 1024, 1)}
