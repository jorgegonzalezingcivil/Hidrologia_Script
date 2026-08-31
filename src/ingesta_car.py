#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Adaptador de ingesta de las series de la CAR.

La Corporación Autónoma Regional de Cundinamarca entrega un libro de Excel en
formato LARGO, una fila por dato, con ocho columnas. Es otro contenedor y otro
vocabulario que el del IDEAM, y por eso vive aparte; pero **entrega al mismo
esquema interno**, de modo que aguas abajo las dos redes son una sola serie
consolidada y ningún módulo tiene que saber de dónde salió cada dato.

QUE ESTE MODULO SI HACE:
    traducir el vocabulario de la CAR al de la cadena, verificar unidades y
    escala, y situar cada estación cruzando con su catálogo.

QUE NO HACE:
    decidir qué estación sirve. El descarte por longitud de serie o por
    consistencia corresponde al M04b y al M05, que lo hacen sobre el dato y
    dejan constancia. Filtrar aquí sería decidir sin evidencia.

Las decisiones de fondo (qué fuente se adopta, qué se descarta y por qué) están
razonadas en docs/especificacion_datos_car.md. El CÓMO se declara en
config/perfiles_car.yaml. Este módulo no contiene ninguna de las dos cosas.
"""

from __future__ import annotations

from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any, Iterator, Sequence

import yaml

from comun.errores import ErrorFormato, ErrorRutas


# El esquema interno es el del M04. Se importa por valor y no por referencia
# para que este módulo se pueda probar sin arrastrar el M04 entero.
CAMPOS_INTERNOS = (
    "codigo", "nombre", "latitud", "longitud", "altitud", "categoria",
    "parametro", "etiqueta", "frecuencia", "fecha", "valor",
    "calificador", "nivel_aprobacion", "fuente",
)

MOTIVO_SIN_EQUIVALENCIA = "parametro y tipo sin equivalencia declarada"
MOTIVO_DESCARTADO = "descarte declarado en el perfil"
MOTIVO_ESCALA = "escala distinta de la esperada"
MOTIVO_UNIDAD = "unidad distinta de la esperada"
MOTIVO_SIN_CATALOGO = "la estacion no esta en el catalogo"
MOTIVO_FECHA = "fecha ilegible"
MOTIVO_VALOR = "valor no numerico"


@dataclass
class Serie:
    """Una equivalencia declarada entre el vocabulario de la CAR y el interno."""

    parametro: str
    tipo: str
    etiqueta: str
    unidad_esperada: str = ""
    nota: str = ""
    con_consumidor: bool = True


@dataclass
class PerfilCar:
    """El formato de entrega de la CAR, declarado en perfiles_car.yaml."""

    fuente: str
    entidad: str
    catalogo: str
    catalogo_crs: str
    columnas: tuple[str, ...]
    campos: dict[str, str]
    formato_fecha: str
    hoja: int
    series: dict[tuple[str, str], Serie]
    parametros_descartados: set[str]
    escala_esperada: str
    exigir_en_catalogo: bool

    def serie_de(self, parametro: str, tipo: str) -> Serie | None:
        return self.series.get((_clave(parametro), _clave(tipo)))


@dataclass
class ResultadoIngesta:
    registros: list[dict[str, Any]] = field(default_factory=list)
    leidos: int = 0
    estaciones: set = field(default_factory=set)
    por_etiqueta: dict[str, int] = field(default_factory=dict)
    descartados: dict[str, int] = field(default_factory=dict)
    ejemplos: dict[str, list] = field(default_factory=dict)
    sin_consumidor: dict[str, int] = field(default_factory=dict)

    def descartar(self, motivo: str, detalle: Any = None) -> None:
        self.descartados[motivo] = self.descartados.get(motivo, 0) + 1
        muestras = self.ejemplos.setdefault(motivo, [])
        if detalle is not None and len(muestras) < 5:
            muestras.append(detalle)


def _clave(texto: Any) -> str:
    """
    Normaliza un rótulo para compararlo: sin espacios sobrantes y en mayúsculas.

    NO se quitan los acentos. 'PRECIPITACIÓN' viene acentuada en el libro y
    acentuada se declara en el perfil; igualarlas por eliminación de tildes
    escondería una discrepancia real entre lo entregado y lo declarado, que es
    justo lo que conviene que salte.
    """
    return " ".join(str(texto or "").split()).upper()


# =============================================================================
# Perfil
# =============================================================================
def cargar_perfil(ruta: str | Path) -> PerfilCar:
    """
    Lee perfiles_car.yaml y devuelve el perfil ya resuelto.

    Excepciones
    -----------
    ErrorRutas
        No existe el archivo.
    ErrorFormato
        Falta una clave obligatoria o una equivalencia está incompleta.
    """
    ruta = Path(ruta)
    if not ruta.is_file():
        raise ErrorRutas(f"no se encuentra el perfil de la CAR en {ruta}.")
    try:
        datos = yaml.safe_load(ruta.read_text(encoding="utf-8")) or {}
    except yaml.YAMLError as error:
        raise ErrorFormato(f"{ruta.name} no es YAML válido: {error}") from error

    series: dict[tuple[str, str], Serie] = {}
    for bloque, con_consumidor in (("series", True), ("series_sin_consumidor", False)):
        for entrada in datos.get(bloque) or []:
            faltan = [c for c in ("parametro", "tipo", "etiqueta")
                      if not entrada.get(c)]
            if faltan:
                raise ErrorFormato(
                    f"una equivalencia de {bloque} en {ruta.name} no declara "
                    f"{', '.join(faltan)}.")
            series[(_clave(entrada["parametro"]), _clave(entrada["tipo"]))] = Serie(
                parametro=entrada["parametro"], tipo=entrada["tipo"],
                etiqueta=entrada["etiqueta"],
                unidad_esperada=entrada.get("unidad_esperada", ""),
                nota=entrada.get("nota", ""), con_consumidor=con_consumidor)

    if not series:
        raise ErrorFormato(f"{ruta.name} no declara ninguna equivalencia.")

    descartados: set[str] = set()
    for entrada in datos.get("descartes") or []:
        if entrada.get("parametro"):
            descartados.add(_clave(entrada["parametro"]))
        for uno in entrada.get("parametros") or []:
            descartados.add(_clave(uno))

    contenedor = datos.get("contenedor") or {}
    controles = datos.get("controles") or {}
    return PerfilCar(
        fuente=datos.get("fuente", "car"),
        entidad=datos.get("entidad", ""),
        catalogo=datos.get("catalogo", ""),
        catalogo_crs=datos.get("catalogo_crs", ""),
        columnas=tuple(datos.get("columnas") or ()),
        campos=dict(datos.get("campos") or {}),
        formato_fecha=datos.get("formato_fecha", ""),
        hoja=int(contenedor.get("hoja", 0)),
        series=series,
        parametros_descartados=descartados,
        escala_esperada=_clave(controles.get("escala_esperada", "")),
        exigir_en_catalogo=bool(controles.get("exigir_en_catalogo", True)),
    )


# =============================================================================
# Lectura del libro
# =============================================================================
def leer_libro(ruta: str | Path, perfil: PerfilCar) -> Iterator[dict[str, Any]]:
    """
    Recorre el libro devolviendo un diccionario por fila.

    Se lee en modo de solo lectura y por filas: el libro entregado tiene 76.248
    filas, y cargarlo entero en memoria como rejilla cuesta sin ninguna ventaja.

    Excepciones
    -----------
    ErrorRutas
        No existe el archivo.
    ErrorFormato
        El encabezado no es el declarado en el perfil. Se comprueba ANTES de
        leer un solo dato: un libro con las columnas cambiadas de sitio se leería
        entero sin protestar y produciría una serie con los campos cruzados.
    """
    import openpyxl

    ruta = Path(ruta)
    if not ruta.is_file():
        raise ErrorRutas(f"no se encuentra el libro de la CAR en {ruta}.")

    libro = openpyxl.load_workbook(ruta, read_only=True, data_only=True)
    try:
        hoja = libro[libro.sheetnames[perfil.hoja]]
        filas = hoja.iter_rows(values_only=True)
        try:
            encabezado = [_clave(c) for c in next(filas)]
        except StopIteration:
            raise ErrorFormato(f"{ruta.name} está vacío.") from None

        esperado = [_clave(c) for c in perfil.columnas]
        if encabezado[:len(esperado)] != esperado:
            raise ErrorFormato(
                f"el encabezado de {ruta.name} no es el declarado en el perfil. "
                f"Esperado {esperado}, encontrado {encabezado[:len(esperado)]}.")

        for fila in filas:
            if fila is None or all(c is None for c in fila):
                continue
            yield dict(zip(encabezado, fila))
    finally:
        libro.close()


def _fecha(valor: Any, formato: str) -> str:
    """Fecha en ISO, o cadena vacía si no es interpretable."""
    if isinstance(valor, datetime):
        return valor.date().isoformat()
    texto = str(valor or "").strip()
    if not texto:
        return ""
    for candidato in (formato, "%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d/%m/%Y"):
        if not candidato:
            continue
        try:
            return datetime.strptime(texto, candidato).date().isoformat()
        except ValueError:
            continue
    return ""


def _numero(valor: Any) -> float | None:
    """Número admitiendo coma decimal. None si no lo es."""
    if isinstance(valor, (int, float)):
        return float(valor)
    texto = str(valor or "").strip().replace(",", ".")
    if not texto:
        return None
    try:
        return float(texto)
    except ValueError:
        return None


# =============================================================================
# Normalización
# =============================================================================
def normalizar(
    fila: dict[str, Any], perfil: PerfilCar,
    catalogo: dict[str, dict[str, Any]] | None = None,
) -> tuple[dict[str, Any] | None, str]:
    """
    Traduce una fila de la CAR al esquema interno de la cadena.

    Devuelve (registro, motivo). Si el registro es None, el motivo dice por qué
    no entró, y quien llama lo cuenta: un descarte sin contar es un descarte que
    nadie puede explicar después.

    SE VERIFICA LA UNIDAD, y no es una formalidad. La columna DATO mezcla
    centímetros de nivel, metros cúbicos por segundo de caudal, milímetros de
    lluvia y grados de temperatura. Una fila de nivel leída como caudal pasaría
    cualquier control numérico sin protestar y contaminaría el análisis de
    crecientes con cifras del orden de los centenares.
    """
    def columna(interno: str, por_defecto: str) -> Any:
        return fila.get(_clave(perfil.campos.get(interno, por_defecto)))

    parametro = _clave(columna("parametro", "PARAMETRO"))
    tipo = _clave(columna("tipo", "TIPO"))
    escala = _clave(columna("escala", "ESCALA"))

    if parametro in perfil.parametros_descartados:
        return None, MOTIVO_DESCARTADO

    serie = perfil.serie_de(parametro, tipo)
    if serie is None:
        return None, MOTIVO_SIN_EQUIVALENCIA

    if perfil.escala_esperada and escala != perfil.escala_esperada:
        return None, MOTIVO_ESCALA

    unidad = _clave(columna("unidades", "UNIDADES"))
    if serie.unidad_esperada and unidad != _clave(serie.unidad_esperada):
        return None, MOTIVO_UNIDAD

    codigo = str(columna("codigo", "CODIGO") or "").strip()
    ficha = (catalogo or {}).get(codigo)
    if perfil.exigir_en_catalogo and ficha is None:
        return None, MOTIVO_SIN_CATALOGO

    fecha = _fecha(columna("fecha", "FECHA"), perfil.formato_fecha)
    if not fecha:
        return None, MOTIVO_FECHA

    valor = _numero(columna("valor", "DATO"))
    if valor is None:
        return None, MOTIVO_VALOR

    ficha = ficha or {}
    registro = {
        "codigo": codigo,
        "nombre": str(columna("nombre", "ESTACION") or "").strip(),
        "latitud": ficha.get("latitud", ""),
        "longitud": ficha.get("longitud", ""),
        "altitud": ficha.get("altitud", ""),
        "categoria": ficha.get("categoria", ""),
        "parametro": serie.parametro,
        "etiqueta": serie.etiqueta,
        "frecuencia": escala.capitalize(),
        "fecha": fecha,
        "valor": valor,
        # La CAR no publica ninguno de los dos. Se dejan vacíos y NO en cero:
        # el M04 usa el nivel de aprobación para decidir cuál conserva ante un
        # conflicto, y un cero se leería como un nivel declarado.
        "calificador": "",
        "nivel_aprobacion": "",
        # De que red salio el dato. Sin esto el M05 no podria distinguir una
        # discrepancia ENTRE REDES de una de una estacion concreta.
        "fuente": perfil.fuente,
    }
    return registro, ""


def ingerir(
    ruta_libro: str | Path, perfil: PerfilCar,
    catalogo: dict[str, dict[str, Any]] | None = None,
) -> ResultadoIngesta:
    """Recorre el libro entero y devuelve los registros ya normalizados."""
    resultado = ResultadoIngesta()
    for fila in leer_libro(ruta_libro, perfil):
        resultado.leidos += 1
        registro, motivo = normalizar(fila, perfil, catalogo)
        if registro is None:
            detalle = (f"{fila.get('CODIGO')} {fila.get('PARAMETRO')} "
                       f"/ {fila.get('TIPO')} [{fila.get('UNIDADES')}]")
            resultado.descartar(motivo, detalle)
            continue
        resultado.registros.append(registro)
        resultado.estaciones.add(registro["codigo"])
        etiqueta = registro["etiqueta"]
        resultado.por_etiqueta[etiqueta] = \
            resultado.por_etiqueta.get(etiqueta, 0) + 1
        serie = perfil.serie_de(_clave(fila.get("PARAMETRO")),
                                _clave(fila.get("TIPO")))
        if serie is not None and not serie.con_consumidor:
            resultado.sin_consumidor[etiqueta] = \
                resultado.sin_consumidor.get(etiqueta, 0) + 1
    return resultado


# =============================================================================
# Catálogo
# =============================================================================
def leer_catalogo(ruta: str | Path, campos: dict[str, str] | None = None,
                  ) -> dict[str, dict[str, Any]]:
    """
    Ficha de cada estación de la CAR, indexada por código.

    EL LIBRO DE DATOS NO TRAE UBICACION. Sin este cruce las estaciones no se
    pueden situar, y por tanto no entran ni a la selección por área ni a
    ninguna interpolación.
    """
    from comun import shapefile

    campos = campos or {
        "codigo": "CODIGO", "nombre": "NOMBRE", "categoria": "CATEGORI_1",
        "altitud": "ELEVACION", "latitud": "N_LATITUD", "longitud": "N_LONGITUD",
    }
    ruta = Path(ruta)
    if not ruta.is_file():
        raise ErrorRutas(f"no se encuentra el catálogo de la CAR en {ruta}.")

    puntos = shapefile.leer_puntos(ruta)
    registros = list(shapefile.leer_registros(ruta))
    if len(puntos) != len(registros):
        raise ErrorFormato(
            f"{ruta.name} tiene {len(puntos)} geometrías y {len(registros)} "
            "registros: no se pueden emparejar.")

    catalogo: dict[str, dict[str, Any]] = {}
    for (lon, lat), fila in zip(puntos, registros):
        codigo = str(fila.get(campos["codigo"], "")).strip()
        if not codigo:
            continue
        catalogo[codigo] = {
            "nombre": str(fila.get(campos["nombre"], "")).strip(),
            "categoria": str(fila.get(campos["categoria"], "")).strip(),
            "altitud": str(fila.get(campos["altitud"], "")).strip(),
            # La COORDENADA sale de la geometría y no de los campos del .dbf:
            # los campos son texto y pueden venir en grados y minutos, mientras
            # que la geometría está siempre en el sistema que declara el .prj.
            "latitud": round(float(lat), 6),
            "longitud": round(float(lon), 6),
        }
    return catalogo


# Campos del catálogo de la CAR. CATEGORI_2 y no CATEGORI_1: el primero trae el
# código corto (PM, PG, CP, CO, LM, LG), que es la nomenclatura del IDEAM; el
# segundo trae el nombre largo, que no sirve para cruzar.
CAMPOS_CATALOGO = {
    "codigo": "CODIGO",
    "nombre": "NOMBRE",
    "categoria": "CATEGORI_2",
    "categoria_desc": "CATEGORI_1",
    "altitud": "ELEVACION",
    "tipo": "TIPO_NOMBR",
    "corriente": "CORRIENT_1",
    "fecha_instalacion": "FECHA_INST",
    "fecha_suspension": "FECHA_SUSP",
}

# Tipos de estación que NO entran. Lo satelital se decide por este campo y no
# por la categoría: hay códigos que delatan el origen (CPS es climatológica
# principal SATELITAL, 54 estaciones), pero fiarse de eso es frágil.
TIPOS_EXCLUIDOS = ("SATELITAL",)


def catalogo_como_ideam(
    ruta: str | Path,
    campos_ideam: dict[str, str],
    entidad: str,
    tipos_excluidos: Sequence[str] = TIPOS_EXCLUIDOS,
) -> tuple[list[dict[str, str]], list[tuple[float, float]], dict[str, int]]:
    """
    El catálogo de la CAR con los NOMBRES DE CAMPO del IDEAM.

    Se traduce en el origen y no en cada consumidor: así el M03 cruza las dos
    redes con el mismo código que ya tenía, y no queda una rama por red que haya
    que mantener en paralelo y que pueda divergir sin que nadie lo note.

    SE EXCLUYE LO SATELITAL. Un valor derivado de píxel no es una medida de
    pluviómetro. Si las 54 climatológicas satelitales entraran como CP, acabarían
    sosteniendo la interpolación de lluvia junto a las estaciones en tierra, que
    es lo que la especificación descarta de forma expresa.

    EL ESTADO NO SE PUEDE TRASLADAR: vale '0' en las 434 estaciones del catálogo
    y no distingue activa de suspendida. Se deja vacío, y el descarte por
    antigüedad lo decide el M04b sobre el dato.

    Devuelve (registros, coordenadas, recuento). Las coordenadas van aparte
    porque salen de la GEOMETRÍA y no del .dbf, igual que en 'leer_catalogo'.

    Excepciones
    -----------
    ErrorRutas
        No existe el catálogo.
    ErrorFormato
        Las geometrías y los registros no se pueden emparejar.
    """
    from comun import shapefile

    ruta = Path(ruta)
    if not ruta.is_file():
        raise ErrorRutas(f"no se encuentra el catálogo de la CAR en {ruta}.")

    puntos = shapefile.leer_puntos(ruta)
    registros = list(shapefile.leer_registros(ruta))
    if len(puntos) != len(registros):
        raise ErrorFormato(
            f"{ruta.name} tiene {len(puntos)} geometrías y {len(registros)} "
            "registros: no se pueden emparejar.")

    excluidos = {_clave(t) for t in tipos_excluidos}
    salida: list[dict[str, str]] = []
    coordenadas: list[tuple[float, float]] = []
    recuento = {"leidas": len(registros), "excluidas_por_tipo": 0,
                "sin_codigo": 0, "admitidas": 0}

    for (lon, lat), fila in zip(puntos, registros):
        codigo = str(fila.get(CAMPOS_CATALOGO["codigo"], "")).strip()
        if not codigo:
            recuento["sin_codigo"] += 1
            continue
        if _clave(fila.get(CAMPOS_CATALOGO["tipo"], "")) in excluidos:
            recuento["excluidas_por_tipo"] += 1
            continue

        traducido: dict[str, str] = {}
        for interno, columna_ideam in campos_ideam.items():
            origen = CAMPOS_CATALOGO.get(interno)
            traducido[columna_ideam] = (
                str(fila.get(origen, "")).strip() if origen else "")

        # La coordenada sale de la geometría, que está en el sistema que declara
        # el .prj; los campos del .dbf pueden venir en grados y minutos.
        if campos_ideam.get("latitud"):
            traducido[campos_ideam["latitud"]] = f"{lat:.6f}"
        if campos_ideam.get("longitud"):
            traducido[campos_ideam["longitud"]] = f"{lon:.6f}"
        # De qué red es la estación. Es lo que permite responder en el informe
        # de dónde salió cada dato, y separar una inconsistencia entre redes de
        # una de una estación concreta.
        if campos_ideam.get("entidad"):
            traducido[campos_ideam["entidad"]] = entidad

        salida.append(traducido)
        coordenadas.append((lon, lat))
        recuento["admitidas"] += 1

    return salida, coordenadas, recuento
